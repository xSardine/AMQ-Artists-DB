import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.styles.alignment import Alignment
from youtubesearchpython import VideosSearch

# ______ General Configuration ______

# Setting to True slow down the process
# but gives you full song link when available on youtube
add_youtube_link = True

# If set to True: will fuse all the json in the folder
# If set to False: will create one sheet per json in the folder
fuse_multiple_json = True

# Sheet styling Configuration
sheet_name = "Sheet1"
link_color = "1155cc"
cell_background_color = "cccccc"
border_color = "949494"
font_police = "Arial"
first_line_font_size = 10
rest_font_size = 10
# Sheet styling Configuration

YT_LINK = "https://www.youtube.com/watch?v="

# ______ General Configuration ______


def song_in_list(song, song_list):

    for song2 in song_list:
        if (
            song["annId"] == song2["annId"]
            and song["Type"] == song2["Type"]
            and song["SongName"] == song2["SongName"]
            and song["Artist"] == song2["Artist"]
        ):
            return True
    return False


def concat(song_list1, song_list2):

    concat_song_list = song_list1

    for song in song_list2:
        if not song_in_list(song, concat_song_list):
            concat_song_list.append(song)

    return concat_song_list


def format_song(song):

    HQlink = song["sept"] if ("sept" in song and song["sept"]) else song["quatre"]
    mp3_link = song["mptrois"] if "mptrois" in song else None
    print(song["SongName"], mp3_link, "\n")

    return {
        "anime_name": song["Anime"],
        "type": song["Type"],
        "info": f"{song['SongName']} by {song['Artist']}",
        "link": HQlink,
        "mp3_link": mp3_link,
    }


def create_workbook(raw_song_list, output_file_name):

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # First line value
    ws.cell(1, 1, "ID")
    ws.cell(1, 2, "Anime Name")
    ws.cell(1, 3, "Song Type")
    ws.cell(1, 4, "Song Info")
    ws.cell(1, 5, "mp3 Links")
    ws.cell(1, 6, "Full Versions")
    ws.cell(1, 7, "Rank")

    # Insert values
    row_iter = 2
    for i, raw_song in enumerate(raw_song_list):

        print(f"Song #{i+1}/{len(raw_song_list)}")
        song = format_song(raw_song)

        yt_link = None

        if add_youtube_link:

            ytsearch = f"{song['info']} full song"
            results = VideosSearch(ytsearch, limit=1).result()

            if not len(results["result"]):
                print(f"New strategy for {song['info']}")
                ytsearch = f"{song['anime_name']} {raw_song['SongName']} full song"
                results = VideosSearch(ytsearch, limit=1).result()

            if len(results["result"]):
                yt_link = YT_LINK + results["result"][0]["id"]

        ws.cell(row_iter, 2, song["anime_name"])
        ws.cell(row_iter, 3, song["type"])
        ws.cell(row_iter, 4, song["info"]).hyperlink = song["link"]
        ws.cell(row_iter, 5, "Link").hyperlink = song["mp3_link"]
        ws.cell(row_iter, 6, "Link").hyperlink = yt_link
        row_iter += 1

    # Change width of column based on longest cell value
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max(
                    (dims.get(cell.column_letter, 0), len(str(cell.value)))
                )
    for col, value in dims.items():
        ws.column_dimensions[col].width = value + 7

    # general style
    gray_color = Color(rgb=cell_background_color)
    gray_background = PatternFill(patternType="solid", fgColor=gray_color)

    for line in ws[f"A1:G{row_iter-1}"]:
        for cell in line:
            cell.fill = gray_background
            cell.font = Font(size=rest_font_size, name=font_police)
            cell.alignment = Alignment(vertical="center")

    # Style for first line
    for line in ws["A1:G1"]:
        for cell in line:
            cell.font = Font(size=first_line_font_size, bold=True, name=font_police)

    # Blue color for links
    for line in ws[f"D2:E{row_iter-1}"]:
        for cell in line:
            cell.font = Font(color=link_color)

    # Sorting property
    ws.auto_filter.ref = f"A1:G{row_iter-1}"

    wb.save(output_file_name)


if __name__ == "__main__":

    json_path = Path(".")
    json_list = list(json_path.glob("*.json"))

    fused_song_list = []
    for json_ in json_list:
        with open(json_, encoding="utf-8") as json_file:
            current_song_list = json.load(json_file)
            if not fuse_multiple_json:
                create_workbook(
                    current_song_list,
                    f'{json_.stem.replace("_", "").replace("SongList", "")}_PR.xlsx',
                )
            else:
                fused_song_list = concat(fused_song_list, current_song_list)

    if fuse_multiple_json:
        create_workbook(fused_song_list, "Fused_PR.xlsx")
